from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import tempfile
import unicodedata
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
BUDGET_SUPABASE_JS = ROOT / "budget-supabase.js"

MONTHS = [
    ("jan", "Janvier", 1),
    ("feb", "Fevrier", 2),
    ("mar", "Mars", 3),
    ("apr", "Avril", 4),
    ("may", "Mai", 5),
    ("jun", "Juin", 6),
    ("jul", "Juillet", 7),
    ("aug", "Aout", 8),
    ("sep", "Septembre", 9),
    ("oct", "Octobre", 10),
    ("nov", "Novembre", 11),
    ("dec", "Decembre", 12),
]
MONTH_BY_KEY = {key: number for key, _label, number in MONTHS}
MONTH_BY_NUMBER = {number: (key, label) for key, label, number in MONTHS}

ENTITY_FILES = [
    {"key": "psa", "label": "PSA", "year": 2026, "file": "activitereelpsa-2026.xlsx"},
    {"key": "gueudet", "label": "Gueudet", "year": 2026, "file": "activitereelgueudet-2026.xlsx"},
    {"key": "ford", "label": "Ford", "year": 2026, "file": "activitereelford-2026.xlsx"},
    {"key": "direct", "label": "Direct", "year": 2026, "file": "activitereeldirect-2026.xlsx"},
    {"key": "psa", "label": "PSA", "year": 2025, "file": "activitereelpsa-2025.xlsx"},
]

CONFIGS = {
    "psa": {
        "client_code": ["n° client interne", "no client interne", "n client interne", "n° client", "numero client interne"],
        "client_name": ["nom du client", "nom client", "client"],
        "amount": ["montant prix achat kent", "montant achat kent", "montant"],
        "month": ["mois2", "mois", "month"],
        "date": ["date facturation", "date de vente", "date vente", "date facture", "date"],
        "reference": ["nos réf kent", "nos ref kent", "reference produits", "référence", "reference", "ref"],
        "designation": ["désignation", "designation", "designation produit", "description produit"],
        "quantity": ["quantité payante servie", "quantite payante servie", "quantité", "quantite", "qte"],
    },
    "gueudet": {
        "client_code": ["n° client interne", "no client interne", "n client interne", "numero client interne", "n° client", "no client", "n client", "numero client", "nclient"],
        "client_name": ["nom du client", "nom client", "client"],
        "amount": ["montant prix achat kent", "montant achat kent", "montant", "ca total", "ca", "chiffre d'affaires", "chiffre daffaires"],
        "month": ["mois2", "mois", "month"],
        "date": ["date facturation", "date de vente", "date vente", "date facture", "date"],
        "reference": ["nos réf kent", "nos ref kent", "reference produits", "référence", "reference", "ref"],
        "designation": ["désignation", "designation", "designation produit", "description produit"],
        "quantity": ["quantité payante servie", "quantite payante servie", "quantité", "quantite", "qte"],
    },
    "default": {
        "client_code": ["code livré", "code livre", "code", "code livre client"],
        "client_name": ["nom client", "nom du client", "client"],
        "amount": ["ca total", "ca", "chiffre d'affaires", "chiffre daffaires", "montant"],
        "month": ["mois2", "mois", "month"],
        "date": ["date facture", "date facturation", "date commande", "date"],
        "reference": ["code produit", "n° produit", "n produit", "reference", "référence", "ref"],
        "designation": ["description produit", "designation", "désignation", "designation produit"],
        "quantity": ["quantité", "quantite", "qte", "qté"],
    },
}


def normalized(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.replace("Â°".lower(), "°")
    return re.sub(r"\s+", " ", text)


def to_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value) if value == value else 0.0
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def cents(value: Any) -> float:
    return round(to_number(value), 2)


def cell_json(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def date_iso(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = str(value or "").strip()
    if not raw:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw[:19], fmt).date().isoformat()
        except ValueError:
            pass
    match = re.match(r"^(\d{1,2})[\/\-.](\d{1,2})[\/\-.](\d{2,4})", raw)
    if match:
        year = int(match.group(3))
        if year < 100:
            year += 2000
        try:
            return date(year, int(match.group(2)), int(match.group(1))).isoformat()
        except ValueError:
            return None
    return None


def month_number(value: Any, fallback_date: Any = None) -> int | None:
    raw = str(value or "").strip()
    number = to_number(raw)
    if 1 <= number <= 12:
        return int(round(number))

    key = normalized(raw).rstrip(".")
    aliases = {
        1: ["janvier", "jan"],
        2: ["fevrier", "fev"],
        3: ["mars", "mar"],
        4: ["avril", "avr"],
        5: ["mai"],
        6: ["juin"],
        7: ["juillet", "juil"],
        8: ["aout", "aou"],
        9: ["septembre", "sept", "sep"],
        10: ["octobre", "oct"],
        11: ["novembre", "nov"],
        12: ["decembre", "dec"],
    }
    for month, values in aliases.items():
        if key in values:
            return month

    iso = date_iso(fallback_date)
    if iso:
        return int(iso[5:7])
    return None


def read_supabase_credentials() -> tuple[str, str]:
    source = BUDGET_SUPABASE_JS.read_text(encoding="utf-8")
    url_match = re.search(r'const SUPABASE_URL = "([^"]+)"', source)
    key_match = re.search(r'const SUPABASE_ANON_KEY = "([^"]+)"', source)
    if not url_match or not key_match:
        raise RuntimeError("Identifiants Supabase introuvables dans budget-supabase.js")
    return url_match.group(1), key_match.group(1)


class SupabaseRest:
    def __init__(self) -> None:
        self.url, self.key = read_supabase_credentials()

    def request(self, method: str, path: str, payload: Any | None = None, prefer: str = "return=representation") -> Any:
        url = f"{self.url}/rest/v1/{path}"
        command = [
            "curl.exe",
            "-k",
            "-sS",
            "-X",
            method,
            url,
            "-H",
            f"apikey: {self.key}",
            "-H",
            f"Authorization: Bearer {self.key}",
            "-H",
            "Content-Type: application/json",
            "-H",
            "Accept: application/json",
            "-w",
            "\n%{http_code}",
        ]
        if prefer:
            command.extend(["-H", f"Prefer: {prefer}"])

        temp_path: str | None = None
        if payload is not None:
            with tempfile.NamedTemporaryFile("w", encoding="utf-8", suffix=".json", delete=False) as tmp:
                json.dump(payload, tmp, ensure_ascii=False)
                temp_path = tmp.name
            command.extend(["--data-binary", f"@{temp_path}"])

        try:
            proc = subprocess.run(command, cwd=ROOT, text=True, capture_output=True, timeout=90)
        finally:
            if temp_path:
                Path(temp_path).unlink(missing_ok=True)

        if proc.returncode != 0:
            raise RuntimeError(f"curl failed: {proc.stderr.strip() or proc.stdout.strip()}")

        raw = proc.stdout
        if "\n" not in raw:
            raise RuntimeError(f"Reponse Supabase illisible: {raw}")
        body, status_text = raw.rsplit("\n", 1)
        status = int(status_text.strip() or "0")
        if status < 200 or status >= 300:
            raise RuntimeError(f"Supabase {method} {path} -> HTTP {status}: {body}")
        return json.loads(body) if body.strip() else None

    def get(self, path: str) -> Any:
        return self.request("GET", path, prefer="")

    def post(self, path: str, payload: Any, prefer: str = "return=representation") -> Any:
        return self.request("POST", path, payload, prefer=prefer)

    def patch(self, path: str, payload: Any, prefer: str = "return=representation") -> Any:
        return self.request("PATCH", path, payload, prefer=prefer)

    def delete(self, path: str, prefer: str = "return=minimal") -> Any:
        return self.request("DELETE", path, prefer=prefer)


def quote_value(value: str) -> str:
    return quote(str(value), safe="")


def get_entity(rest: SupabaseRest, key: str) -> dict[str, Any]:
    rows = rest.get(f"budget_entites?select=*&key=eq.{quote_value(key)}&limit=1")
    if not rows:
        raise RuntimeError(f"Entite introuvable dans Supabase: {key}")
    return rows[0]


def find_col(header_map: dict[str, int], candidates: list[str]) -> int | None:
    for candidate in candidates:
        idx = header_map.get(normalized(candidate))
        if idx is not None:
            return idx
    return None


def best_sheet(workbook: Any) -> Any:
    return max(workbook.worksheets, key=lambda ws: (ws.max_row or 0) * (ws.max_column or 0))


def parse_file(spec: dict[str, Any]) -> dict[str, Any]:
    path = ROOT / spec["file"]
    if not path.exists():
        raise FileNotFoundError(path)

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = best_sheet(wb)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError(f"Fichier vide: {path.name}")

    headers = [cell_json(value) for value in rows[0]]
    header_map = {normalized(value): idx for idx, value in enumerate(headers) if normalized(value)}
    config = CONFIGS.get(spec["key"], CONFIGS["default"])

    col_client_code = find_col(header_map, config["client_code"])
    col_client_name = find_col(header_map, config["client_name"])
    col_amount = find_col(header_map, config["amount"])
    col_month = find_col(header_map, config["month"])
    col_date = find_col(header_map, config["date"])
    col_ref = find_col(header_map, config["reference"])
    col_designation = find_col(header_map, config["designation"])
    col_quantity = find_col(header_map, config["quantity"])

    missing = []
    if col_client_code is None:
        missing.append("code client")
    if col_client_name is None:
        missing.append("nom client")
    if col_amount is None:
        missing.append("montant")
    if col_month is None and col_date is None:
        missing.append("mois/date")
    if missing:
        raise RuntimeError(f"{path.name}: colonnes manquantes: {', '.join(missing)}")

    month_groups: dict[int, list[dict[str, Any]]] = {}
    skipped = {"empty": 0, "no_month": 0, "no_amount": 0}

    for order, row in enumerate(rows[1:], start=1):
        client_code = str(row[col_client_code] or "").strip() if col_client_code is not None else ""
        client_name = str(row[col_client_name] or "").strip() if col_client_name is not None else ""
        if not client_code and not client_name:
            skipped["empty"] += 1
            continue

        date_value = row[col_date] if col_date is not None else None
        month_value = row[col_month] if col_month is not None else None
        month = month_number(month_value, date_value)
        if not month:
            skipped["no_month"] += 1
            continue

        amount = cents(row[col_amount]) if col_amount is not None else 0.0
        if not amount:
            skipped["no_amount"] += 1
            continue

        raw_data = {str(headers[idx] or f"col_{idx+1}"): cell_json(row[idx] if idx < len(row) else "") for idx in range(len(headers))}
        line = {
            "ordre": order,
            "client_code": client_code or None,
            "client_nom": client_name or f"Client {client_code}",
            "montant": amount,
            "mois_source": month,
            "date_piece": date_iso(date_value),
            "reference": str(row[col_ref] or "").strip() if col_ref is not None else None,
            "designation": str(row[col_designation] or "").strip() if col_designation is not None else None,
            "quantite": to_number(row[col_quantity]) if col_quantity is not None else 0.0,
            "raw_data": raw_data,
        }
        month_groups.setdefault(month, []).append(line)

    columns = {
        "clientCode": headers[col_client_code] if col_client_code is not None else None,
        "clientName": headers[col_client_name] if col_client_name is not None else None,
        "amount": headers[col_amount] if col_amount is not None else None,
        "month": headers[col_month] if col_month is not None else None,
        "date": headers[col_date] if col_date is not None else None,
        "reference": headers[col_ref] if col_ref is not None else None,
        "designation": headers[col_designation] if col_designation is not None else None,
        "quantity": headers[col_quantity] if col_quantity is not None else None,
    }
    return {
        "spec": spec,
        "sheet": ws.title,
        "columns": columns,
        "months": month_groups,
        "skipped": skipped,
    }


def chunked(values: list[dict[str, Any]], size: int = 500) -> list[list[dict[str, Any]]]:
    return [values[index : index + size] for index in range(0, len(values), size)]


def active_imports(rest: SupabaseRest, entity_id: str, year: int, month: int) -> list[dict[str, Any]]:
    return rest.get(
        "reel_imports"
        f"?select=*&entite_id=eq.{quote_value(entity_id)}&annee=eq.{year}&mois=eq.{month}&statut=eq.active"
    ) or []


def deactivate_imports(rest: SupabaseRest, import_ids: list[str], execute: bool) -> None:
    if not import_ids or not execute:
        return
    rest.patch(
        "reel_imports?id=in.(" + ",".join(quote_value(value) for value in import_ids) + ")",
        {"statut": "inactive"},
        prefer="return=minimal",
    )


def create_month_import(rest: SupabaseRest, entity: dict[str, Any], parsed: dict[str, Any], month: int, lines: list[dict[str, Any]], execute: bool) -> dict[str, Any]:
    spec = parsed["spec"]
    _key, month_label = MONTH_BY_NUMBER[month]
    year = int(spec["year"])
    total = cents(sum(line["montant"] for line in lines))
    existing = active_imports(rest, entity["id"], year, month)
    summary = {
        "entity": spec["label"],
        "year": year,
        "month": month,
        "month_label": month_label,
        "lines": len(lines),
        "total": total,
        "existing_active": len(existing),
    }
    if not execute:
        return summary

    replaced_ids = [row["id"] for row in existing]
    deactivate_imports(rest, replaced_ids, execute=True)

    now = datetime.now(timezone.utc).isoformat()
    import_payload = {
        "entite_id": entity["id"],
        "annee": year,
        "mois": month,
        "statut": "active",
        "nom": f"Reel {entity['libelle']} {month_label} {year} - import Excel",
        "source_file": spec["file"],
        "sheet_name": parsed["sheet"],
        "total_mois": total,
        "nb_lignes": len(lines),
        "colonnes_map": parsed["columns"],
        "meta": {
            "source": "import_excel_reels_local",
            "imported_at": now,
            "replaced_import_ids": replaced_ids,
            "skipped": parsed["skipped"],
        },
    }

    saved = None
    try:
        saved = rest.post("reel_imports", import_payload)[0]
        for chunk in chunked([{**line, "import_id": saved["id"]} for line in lines]):
            rest.post("reel_lignes", chunk, prefer="return=minimal")
        summary["import_id"] = saved["id"]
        return summary
    except Exception:
        if saved and saved.get("id"):
            rest.delete(f"reel_imports?id=eq.{quote_value(saved['id'])}")
        if replaced_ids:
            rest.patch(
                "reel_imports?id=in.(" + ",".join(quote_value(value) for value in replaced_ids) + ")",
                {"statut": "active"},
                prefer="return=minimal",
            )
        raise


def run(execute: bool, only_year: int | None = None) -> dict[str, Any]:
    rest = SupabaseRest()
    results = []
    for spec in ENTITY_FILES:
        if only_year and int(spec["year"]) != only_year:
            continue
        path = ROOT / spec["file"]
        if not path.exists():
            results.append({"entity": spec["label"], "year": spec["year"], "file": spec["file"], "missing": True})
            continue
        entity = get_entity(rest, spec["key"])
        parsed = parse_file(spec)
        for month, lines in sorted(parsed["months"].items()):
            results.append(create_month_import(rest, entity, parsed, month, lines, execute))
    return {"execute": execute, "results": results}


def main() -> int:
    parser = argparse.ArgumentParser(description="Importe les fichiers Excel de reel mensuel dans Supabase.")
    parser.add_argument("--execute", action="store_true", help="Ecrit dans Supabase. Sans ce flag, affiche seulement l'aperçu.")
    parser.add_argument("--year", type=int, default=None, help="Limiter a une annee.")
    args = parser.parse_args()

    result = run(execute=args.execute, only_year=args.year)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
