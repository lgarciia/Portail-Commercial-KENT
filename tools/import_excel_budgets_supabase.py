from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
BUDGET_SUPABASE_JS = ROOT / "budget-supabase.js"
YEAR = 2026

MONTHS = [
    ("jan", "janvier"),
    ("feb", "fevrier"),
    ("mar", "mars"),
    ("apr", "avril"),
    ("may", "mai"),
    ("jun", "juin"),
    ("jul", "juillet"),
    ("aug", "aout"),
    ("sep", "septembre"),
    ("oct", "octobre"),
    ("nov", "novembre"),
    ("dec", "decembre"),
]

ENTITIES = [
    {"key": "psa", "label": "PSA", "file": "budgetpsa-2026.xlsx"},
    {"key": "gueudet", "label": "Gueudet", "file": "budgetgueudet-2026.xlsx"},
    {"key": "ford", "label": "Ford", "file": "budgetford-2026.xlsx"},
    {"key": "direct", "label": "Direct", "file": "budgetdirect-2026.xlsx"},
]


def normalized(value: Any) -> str:
    import unicodedata

    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"\s+", " ", text)


def to_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value) if value == value else 0.0
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def cents(value: Any) -> float:
    return round(to_number(value), 2)


def read_supabase_credentials() -> tuple[str, str]:
    source = BUDGET_SUPABASE_JS.read_text(encoding="utf-8")
    url_match = re.search(r'const SUPABASE_URL = "([^"]+)"', source)
    key_match = re.search(r'const SUPABASE_ANON_KEY = "([^"]+)"', source)
    if not url_match or not key_match:
        raise RuntimeError("Identifiants Supabase introuvables dans budget-supabase.js")
    return url_match.group(1), key_match.group(1)


class SupabaseRest:
    def __init__(self) -> None:
        self.url, self.key = read_supabase_credentials()

    def request(self, method: str, path: str, payload: Any | None = None, prefer: str = "return=representation") -> Any:
        url = f"{self.url}/rest/v1/{path}"
        command = [
            "curl.exe",
            "-k",
            "-sS",
            "-X",
            method,
            url,
            "-H",
            f"apikey: {self.key}",
            "-H",
            f"Authorization: Bearer {self.key}",
            "-H",
            "Content-Type: application/json",
            "-H",
            "Accept: application/json",
            "-w",
            "\n%{http_code}",
        ]
        if prefer:
            command.extend(["-H", f"Prefer: {prefer}"])

        temp_path: str | None = None
        if payload is not None:
            with tempfile.NamedTemporaryFile("w", encoding="utf-8", suffix=".json", delete=False) as tmp:
                json.dump(payload, tmp, ensure_ascii=False)
                temp_path = tmp.name
            command.extend(["--data-binary", f"@{temp_path}"])

        try:
            proc = subprocess.run(command, cwd=ROOT, text=True, capture_output=True, timeout=60)
        finally:
            if temp_path:
                Path(temp_path).unlink(missing_ok=True)

        if proc.returncode != 0:
            raise RuntimeError(f"curl failed: {proc.stderr.strip() or proc.stdout.strip()}")

        raw = proc.stdout
        if "\n" not in raw:
            raise RuntimeError(f"Reponse Supabase illisible: {raw}")
        body, status_text = raw.rsplit("\n", 1)
        status = int(status_text.strip() or "0")
        if status < 200 or status >= 300:
            raise RuntimeError(f"Supabase {method} {path} -> HTTP {status}: {body}")
        return json.loads(body) if body.strip() else None

    def get(self, path: str) -> Any:
        return self.request("GET", path, prefer="")

    def post(self, path: str, payload: Any, prefer: str = "return=representation") -> Any:
        return self.request("POST", path, payload, prefer=prefer)

    def patch(self, path: str, payload: Any, prefer: str = "return=representation") -> Any:
        return self.request("PATCH", path, payload, prefer=prefer)


def find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(rows):
        header_map = {normalized(cell): idx for idx, cell in enumerate(row) if normalized(cell)}
        if "client" in header_map and any(label in header_map for _, label in MONTHS):
            return row_index, header_map
    raise RuntimeError("En-tetes budget introuvables")


def parse_budget_file(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx, header_map = find_header_row(rows[:20])

    client_col = header_map["client"]
    number_col = header_map.get("n client") or header_map.get("n° client") or header_map.get("numero client") or header_map.get("num client")
    comment_col = header_map.get("commentaire") or header_map.get("comment")
    total_col = header_map.get("total")
    month_cols = {key: header_map[label] for key, label in MONTHS}

    lines: list[dict[str, Any]] = []
    total_months = 0.0
    total_excel_col = 0.0

    for order, row in enumerate(rows[header_idx + 1 :], start=1):
        client_name = str(row[client_col] or "").strip()
        if not client_name:
            continue
        line: dict[str, Any] = {
            "ordre": len(lines) + 1,
            "client_nom": client_name,
            "numero_client": str(row[number_col] or "").strip() if number_col is not None else None,
            "commentaire": str(row[comment_col] or "").strip() if comment_col is not None else None,
        }
        line_total = 0.0
        for key, _label in MONTHS:
            amount = cents(row[month_cols[key]])
            line[key] = amount
            line_total += amount
        line["total"] = cents(line_total)
        lines.append(line)
        total_months += line["total"]
        if total_col is not None:
            total_excel_col += cents(row[total_col])

    return {
        "file": path.name,
        "lines": lines,
        "line_count": len(lines),
        "total_months": cents(total_months),
        "total_excel_column": cents(total_excel_col),
    }


def quote_value(value: str) -> str:
    return quote(value, safe="")


def get_entity(rest: SupabaseRest, key: str) -> dict[str, Any] | None:
    rows = rest.get(f"budget_entites?select=*&key=eq.{quote_value(key)}&limit=1")
    return rows[0] if rows else None


def get_active_budget(rest: SupabaseRest, entity_id: str, year: int) -> dict[str, Any] | None:
    rows = rest.get(
        "budgets"
        f"?select=*&entite_id=eq.{quote_value(entity_id)}&annee=eq.{year}&statut=eq.active&limit=1"
    )
    return rows[0] if rows else None


def get_budget_lines(rest: SupabaseRest, budget_id: str) -> list[dict[str, Any]]:
    return rest.get(
        "budget_lignes"
        f"?select=*&budget_id=eq.{quote_value(budget_id)}&order=ordre.asc"
    ) or []


def summarize_lines(lines: list[dict[str, Any]]) -> dict[str, Any]:
    total = 0.0
    for line in lines:
        total += sum(cents(line.get(key)) for key, _label in MONTHS)
    return {"line_count": len(lines), "total_months": cents(total)}


def deactivate_budget(rest: SupabaseRest, budget_id: str, execute: bool) -> dict[str, Any]:
    summary = {"deactivate_action": "deactivate_existing_active", "deactivated_budget_id": budget_id}
    if not execute:
        return summary
    payload = {
        "statut": "inactive",
        "validated_at": None,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    rest.patch(f"budgets?id=eq.{quote_value(budget_id)}", payload, prefer="return=minimal")
    return summary


def create_budget(
    rest: SupabaseRest,
    entity: dict[str, Any],
    parsed: dict[str, Any],
    execute: bool,
    replaced_budget_id: str | None = None,
) -> dict[str, Any]:
    name = f"Budget {entity['libelle']} {YEAR} - import Excel"
    summary = {
        "action": "create_active_budget",
        "budget_name": name,
        "line_count": parsed["line_count"],
        "total_months": parsed["total_months"],
    }
    if not execute:
        return summary

    now = datetime.now(timezone.utc).isoformat()
    budget_payload = {
        "entite_id": entity["id"],
        "nom": name,
        "annee": YEAR,
        "statut": "active",
        "total_annuel": parsed["total_months"],
        "nb_lignes": parsed["line_count"],
        "validated_at": now,
        "meta": {
            "source": "import_excel_local",
            "file": parsed["file"],
            "imported_at": now,
            "mode": "safe_missing_active_only",
        },
    }
    if replaced_budget_id:
        budget_payload["meta"]["replaced_budget_id"] = replaced_budget_id
        budget_payload["meta"]["mode"] = "replace_existing_active_without_delete"
    budget = rest.post("budgets", budget_payload)[0]

    lines = []
    for line in parsed["lines"]:
        payload = dict(line)
        payload["budget_id"] = budget["id"]
        lines.append(payload)

    batch_size = 100
    for start in range(0, len(lines), batch_size):
        rest.post("budget_lignes", lines[start : start + batch_size], prefer="return=minimal")

    summary["budget_id"] = budget["id"]
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Importe les budgets Excel 2026 manquants dans Supabase.")
    parser.add_argument("--execute", action="store_true", help="Cree les budgets actifs manquants.")
    parser.add_argument(
        "--replace-entity",
        choices=[entity["key"] for entity in ENTITIES],
        help="Desactive le budget actif de cette entite puis cree un nouveau budget actif depuis Excel, sans supprimer l'ancien.",
    )
    args = parser.parse_args()

    rest = SupabaseRest()
    report: list[dict[str, Any]] = []

    for config in ENTITIES:
        file_path = ROOT / config["file"]
        if not file_path.exists():
            report.append({"entity": config["key"], "file": config["file"], "status": "missing_file"})
            continue

        parsed = parse_budget_file(file_path)
        entity = get_entity(rest, config["key"])
        if not entity:
            report.append({"entity": config["key"], "file": config["file"], "status": "missing_entity"})
            continue

        active = get_active_budget(rest, entity["id"], YEAR)
        item: dict[str, Any] = {
            "entity": config["key"],
            "label": config["label"],
            "file": config["file"],
            "excel_lines": parsed["line_count"],
            "excel_total_months": parsed["total_months"],
            "excel_total_column": parsed["total_excel_column"],
        }

        if active:
            if args.replace_entity == config["key"]:
                existing_lines = get_budget_lines(rest, active["id"])
                existing_summary = summarize_lines(existing_lines)
                item.update(
                    {
                        "status": "replaced" if args.execute else "would_replace",
                        "old_active_budget": active["nom"],
                        "old_active_budget_id": active["id"],
                        "old_supabase_lines": existing_summary["line_count"],
                        "old_supabase_total_months": existing_summary["total_months"],
                    }
                )
                item.update(deactivate_budget(rest, active["id"], execute=args.execute))
                item.update(create_budget(rest, entity, parsed, execute=args.execute, replaced_budget_id=active["id"]))
                report.append(item)
                continue

            existing_lines = get_budget_lines(rest, active["id"])
            existing_summary = summarize_lines(existing_lines)
            item.update(
                {
                    "status": "skipped_existing_active",
                    "active_budget": active["nom"],
                    "active_budget_id": active["id"],
                    "supabase_lines": existing_summary["line_count"],
                    "supabase_total_months": existing_summary["total_months"],
                    "delta_excel_vs_supabase": cents(parsed["total_months"] - existing_summary["total_months"]),
                }
            )
        else:
            item.update({"status": "created" if args.execute else "would_create"})
            item.update(create_budget(rest, entity, parsed, execute=args.execute))

        report.append(item)

    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
